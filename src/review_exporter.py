"""
review_exporter.py
Exports classified and validated rows to output/review_rows.xlsx.
Columns are ordered for human review. Validation warnings appear as a final column.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Human-readable column order for review
REVIEW_COLUMNS = [
    "Title",
    "ProblemPainPoint",
    "EvidenceSummary",
    "SourceSpeaker",
    "SourceTimestamp",
    "ConfidenceLevel",           # Column F
    "MaturitySignal",            # Column G
    "AIUseCaseType",
    "PrimaryFunctionChoice",
    "LevelOfAnalysis",
    "OperatingBucket",
    "WorkstreamType",
    "ProcessStage",              # Column L
    "SubOrdinateFunction",       # Column M
    "UpstreamDownstreamImpact",
    "SignalStrength",
    "RequestingTeam",
    "SuggestedBusinessOwnerText",
    "SuggestedTechnicalOwnerText",
    "SuggestedSMEChampionText",
    "PrimaryTool",
    "PrimaryDataSource",
    "DataSensitivity",
    "AutomationRisk",
    "GuardrailsNeeded",
    "SecurityAccessConcern",
    "LegalComplianceConcern",
    "HumanInTheLoopRequired",
    "HumanReviewRequired",
    "IntegrationNeeded",
    "FrequencyOfPainPoint",
    "ManualEffortLevel",
    "Repeatability",
    "ScalabilityPotential",
    "NextStep",
    "CurrentStatus",
    "Priority",
    "ScheduleHealth",
    "ValueScore",
    "EffortScore",
    "RiskScore",
    "ReadinessScore",
    "SignalScore",
    "review_status",
    "_validation_warnings",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_ROW_FILL = PatternFill("solid", fgColor="EEF2F7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
APPROVED_FILL = PatternFill("solid", fgColor="D5F5E3")
REJECTED_FILL = PatternFill("solid", fgColor="FADBD8")
PENDING_FILL = PatternFill("solid", fgColor="FEF9E7")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_cell_value(value) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if value is None:
        return ""
    return str(value)


def export_review_workbook(
    rows: list[dict],
    output_path: str = "output/review_rows.xlsx",
    triage_rows: list[dict] | None = None,
) -> None:
    """
    Write classified rows to a formatted Excel workbook.
    Primary rows go to the main "Opportunities" sheet.
    triage_rows (low confidence / session cap excess) go to a second "Triage" sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    # Write header row
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 36

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        has_warnings = bool(row_data.get("_validation_warnings"))
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            raw = row_data.get(col_name)
            if col_name == "_validation_warnings" and isinstance(raw, list):
                value = " | ".join(raw)
            else:
                value = _format_cell_value(raw)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if col_name == "_validation_warnings" and value:
                cell.fill = WARN_FILL
            elif col_name == "review_status":
                if value == "approved":
                    cell.fill = APPROVED_FILL
                elif value == "rejected":
                    cell.fill = REJECTED_FILL
                elif not value:
                    cell.fill = PENDING_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-fit column widths (capped)
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        max_len = len(col_name)
        for row_data in rows:
            val = _format_cell_value(row_data.get(col_name))
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

    ws.freeze_panes = "A2"

    # ── Triage sheet ─────────────────────────────────────────────────────────
    if triage_rows:
        TRIAGE_HEADER_FILL = PatternFill("solid", fgColor="7B3F00")
        wt = wb.create_sheet(title="Triage")
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = wt.cell(row=1, column=col_idx, value=col_name)
            cell.fill = TRIAGE_HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        wt.row_dimensions[1].height = 36

        for row_idx, row_data in enumerate(triage_rows, start=2):
            for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
                raw = row_data.get(col_name)
                value = " | ".join(raw) if col_name == "_validation_warnings" and isinstance(raw, list) \
                    else _format_cell_value(raw)
                cell = wt.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            max_len = len(col_name)
            for row_data in triage_rows:
                val = _format_cell_value(row_data.get(col_name))
                max_len = max(max_len, min(len(val), 60))
            wt.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
        wt.freeze_panes = "A2"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    triage_note = f" + {len(triage_rows)} triage" if triage_rows else ""
    print(f"[exporter] {len(rows)} rows{triage_note} -> {output_path}")
