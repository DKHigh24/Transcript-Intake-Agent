"""
master_exporter.py

Maintains output/master_opportunities.xlsx — a flat, append-mode workbook
that accumulates every classified opportunity across all weekly runs.

Designed for Power BI consumption:
  - One row per opportunity per week it was identified
  - Column K = ProcessStage, Column L = SubOrdinateFunction (mirrors review_rows.xlsx)
  - WeekDate column (last) for time-slicing in Power BI
  - Idempotent: re-running the same week replaces that week's rows rather than duplicating

Call update_master(rows, meeting_date) from main.py after each weekly run.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER_PATH = Path("output/master_opportunities.xlsx")

# Column order mirrors review_rows.xlsx with SubOrdinateFunction at column L.
# WeekDate is appended as the final column for Power BI time filtering.
MASTER_COLUMNS = [
    "Title",                        # A
    "ProblemPainPoint",             # B
    "EvidenceSummary",              # C
    "SourceSpeaker",                # D
    "SourceTimestamp",              # E
    "ConfidenceLevel",              # F
    "MaturitySignal",               # G
    "AIUseCaseType",                # H
    "PrimaryFunctionChoice",        # I
    "LevelOfAnalysis",              # J
    "OperatingBucket",              # K
    "ProcessStage",                 # L
    "SubOrdinateFunction",          # M
    "UpstreamDownstreamImpact",     # N
    "SignalStrength",               # O
    "RequestingTeam",               # P
    "SuggestedBusinessOwnerText",   # Q
    "SuggestedTechnicalOwnerText",  # R
    "SuggestedSMEChampionText",     # S
    "PrimaryTool",                  # T
    "PrimaryDataSource",            # U
    "DataSensitivity",              # V
    "AutomationRisk",               # W
    "GuardrailsNeeded",             # X
    "SecurityAccessConcern",        # Y
    "LegalComplianceConcern",       # Z
    "HumanInTheLoopRequired",       # AA
    "HumanReviewRequired",          # AB
    "IntegrationNeeded",            # AC
    "FrequencyOfPainPoint",         # AD
    "ManualEffortLevel",            # AE
    "Repeatability",                # AF
    "ScalabilityPotential",         # AG
    "NextStep",                     # AH
    "CurrentStatus",                # AI
    "Priority",                     # AJ
    "ScheduleHealth",               # AK
    "ValueScore",                   # AL
    "EffortScore",                  # AM
    "RiskScore",                    # AN
    "ReadinessScore",               # AO
    "SignalScore",                  # AP
    "WeekDate",                     # AQ — Power BI time slicer
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_ROW_FILL = PatternFill("solid", fgColor="EEF2F7")
WEEK_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _fmt(value) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if value is None:
        return ""
    return str(value)


def _write_header(ws) -> None:
    for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def _write_rows(ws, rows: list[dict], start_row: int) -> None:
    for row_idx, row_data in enumerate(rows, start=start_row):
        for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
            value = _fmt(row_data.get(col_name))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if col_name == "WeekDate":
                cell.fill = WEEK_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL


def _set_col_widths(ws, all_rows: list[dict]) -> None:
    for col_idx, col_name in enumerate(MASTER_COLUMNS, start=1):
        max_len = len(col_name)
        for row_data in all_rows:
            val = _fmt(row_data.get(col_name))
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def update_master(rows: list[dict], meeting_date: str) -> Path:
    """
    Append `rows` for `meeting_date` to the master workbook.
    If rows for this week already exist they are replaced (idempotent).
    Returns the path to the master file.
    """
    # Stamp each row with the week date
    stamped = [{**r, "WeekDate": meeting_date} for r in rows]

    if MASTER_PATH.exists():
        wb = load_workbook(str(MASTER_PATH))
        ws = wb.active

        # Collect all existing rows except those belonging to this week
        existing: list[dict] = []
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(MASTER_COLUMNS) + 1)]
        for row_idx in range(2, ws.max_row + 1):
            row_vals = {headers[c]: ws.cell(row=row_idx, column=c + 1).value for c in range(len(headers))}
            if row_vals.get("WeekDate") != meeting_date:
                existing.append(row_vals)

        all_rows = existing + stamped

        # Rebuild the sheet cleanly
        ws.delete_rows(1, ws.max_row)
        _write_header(ws)
        _write_rows(ws, all_rows, start_row=2)
        _set_col_widths(ws, all_rows)
    else:
        MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Opportunities"
        _write_header(ws)
        _write_rows(ws, stamped, start_row=2)
        _set_col_widths(ws, stamped)

    wb.save(str(MASTER_PATH))
    total = (ws.max_row - 1) if ws.max_row > 1 else len(stamped)
    print(f"[master] {len(stamped)} rows added for {meeting_date} -> {MASTER_PATH} ({total} total rows)")
    return MASTER_PATH


def rebuild_master() -> Path:
    """
    Rebuild master_opportunities.xlsx from scratch using all archived
    classified_rows.json files.  Useful after a back-fill operation.
    """
    import json
    from pathlib import Path as _Path

    weeks_dir = _Path("output/weeks")
    week_dirs = sorted(weeks_dir.iterdir()) if weeks_dir.exists() else []

    if MASTER_PATH.exists():
        MASTER_PATH.unlink()

    for week_dir in week_dirs:
        rows_file = week_dir / "classified_rows.json"
        if not rows_file.exists():
            continue
        rows = json.loads(rows_file.read_text(encoding="utf-8"))
        update_master(rows, week_dir.name)

    print(f"[master] rebuild complete -> {MASTER_PATH}")
    return MASTER_PATH
